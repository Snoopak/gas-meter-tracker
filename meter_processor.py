#!/usr/bin/env python3
"""
Система обробки фотографій лічильників з автоматичним розпізнаванням даних
"""

import cv2
import pytesseract
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
from datetime import datetime
from pathlib import Path


class MeterProcessor:
    def __init__(self, excel_file='lichilnyky.xlsx'):
        self.excel_file = excel_file
        self.setup_excel()
    
    def setup_excel(self):
        """Створює або завантажує Excel файл з налаштуваннями"""
        if os.path.exists(self.excel_file):
            self.wb = load_workbook(self.excel_file)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = 'Лічильники'
            
            # Заголовки таблиці
            headers = ['№', 'Дата', 'Марка', 'Серійний номер', 'Рік випуску', 'Показання', 'Фото']
            self.ws.append(headers)
            
            # Форматування заголовків
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, header in enumerate(headers, 1):
                cell = self.ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ширина колонок
            self.ws.column_dimensions['A'].width = 8
            self.ws.column_dimensions['B'].width = 15
            self.ws.column_dimensions['C'].width = 20
            self.ws.column_dimensions['D'].width = 20
            self.ws.column_dimensions['E'].width = 12
            self.ws.column_dimensions['F'].width = 15
            self.ws.column_dimensions['G'].width = 30
            
            self.wb.save(self.excel_file)
    
    def preprocess_image(self, image_path):
        """Попередня обробка зображення для кращого розпізнавання"""
        img = cv2.imread(image_path)
        if img is None:
            raise ValueError(f"Не вдалося завантажити зображення: {image_path}")
        
        # Конвертація в градації сірого
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        # Покращення контрасту
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(gray)
        
        # Бінаризація
        _, binary = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        
        # Зменшення шуму
        denoised = cv2.fastNlMeansDenoising(binary, None, 10, 7, 21)
        
        return img, denoised
    
    def extract_text_from_image(self, image_path):
        """Розпізнає текст з фото за допомогою OCR"""
        try:
            _, processed = self.preprocess_image(image_path)
            
            # Налаштування Tesseract для української та англійської мов
            custom_config = r'--oem 3 --psm 6 -l ukr+eng'
            text = pytesseract.image_to_string(processed, config=custom_config)
            
            return text
        except Exception as e:
            print(f"Помилка розпізнавання тексту: {e}")
            return ""
    
    def parse_meter_data(self, text, image_path):
        """Витягує дані про лічильник з розпізнаного тексту"""
        data = {
            'date': datetime.now().strftime('%d.%m.%Y'),
            'brand': '',
            'serial_number': '',
            'year': '',
            'reading': '',
            'photo': os.path.basename(image_path)
        }
        
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        # Пошук серійного номера (різні формати)
        serial_patterns = [
            r'[Сс]ер[іi]йний\s*[№#]?\s*:?\s*([A-Z0-9\-]+)',
            r'[Ss]erial\s*[Nn]o?\s*:?\s*([A-Z0-9\-]+)',
            r'[№#]\s*([A-Z0-9\-]{5,})',
            r'\b([A-Z]{2,}\d{4,})\b',
            r'\b(\d{8,})\b'
        ]
        
        for pattern in serial_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                data['serial_number'] = match.group(1)
                break
        
        # Пошук року випуску
        year_patterns = [
            r'[Рр]ік\s+вип[^\d]*(\d{4})',
            r'[Yy]ear\s*:?\s*(\d{4})',
            r'\b(19\d{2}|20[0-2]\d)\b'
        ]
        
        for pattern in year_patterns:
            match = re.search(pattern, text)
            if match:
                year = match.group(1)
                if 1990 <= int(year) <= 2025:
                    data['year'] = year
                    break
        
        # Пошук показань (числа з можливими пробілами або крапками)
        reading_patterns = [
            r'[Пп]оказання\s*:?\s*([\d\s.]+)',
            r'[Rr]eading\s*:?\s*([\d\s.]+)',
            r'\b(\d{1,6}(?:\.\d{1,3})?)\s*(?:кВт|m³|м³)',
            r'^[\s]*(\d{4,7})[\s]*$'  # Окреме число на рядку
        ]
        
        for pattern in reading_patterns:
            match = re.search(pattern, text, re.MULTILINE)
            if match:
                reading = match.group(1).replace(' ', '').strip()
                if reading and reading.replace('.', '').isdigit():
                    data['reading'] = reading
                    break
        
        # Пошук марки (перші слова або слова перед серійним номером)
        brand_patterns = [
            r'[Мм]арка\s*:?\s*([A-ZА-ЯІЇЄa-zа-яіїє\s\-]+?)(?:\n|[Сс]ер)',
            r'^([A-ZА-ЯІЇЄa-z][A-ZА-ЯІЇЄa-zа-яіїє\s\-]{2,20})(?:\n|\s{2,})',
        ]
        
        for pattern in brand_patterns:
            match = re.search(pattern, text, re.MULTILINE | re.IGNORECASE)
            if match:
                brand = match.group(1).strip()
                if len(brand) > 2 and not brand.isdigit():
                    data['brand'] = brand
                    break
        
        # Якщо марку не знайдено, беремо перше слово
        if not data['brand'] and lines:
            first_word = lines[0].split()[0] if lines[0].split() else ''
            if first_word and not first_word.isdigit() and len(first_word) > 2:
                data['brand'] = first_word
        
        return data
    
    def add_to_excel(self, data):
        """Додає дані в Excel таблицю"""
        # Визначаємо номер запису
        row_num = self.ws.max_row + 1
        record_num = row_num - 1
        
        # Додаємо дані
        row_data = [
            record_num,
            data['date'],
            data['brand'],
            data['serial_number'],
            data['year'],
            data['reading'],
            data['photo']
        ]
        
        self.ws.append(row_data)
        
        # Форматування рядка
        alignment = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num in range(1, 8):
            cell = self.ws.cell(row=row_num, column=col_num)
            cell.alignment = alignment
            cell.border = border
        
        # Зберігаємо файл
        self.wb.save(self.excel_file)
        print(f"✓ Дані додано в Excel (рядок {row_num})")
    
    def process_image(self, image_path, manual_data=None):
        """Обробляє фото лічильника"""
        print(f"\n{'='*60}")
        print(f"Обробка фото: {os.path.basename(image_path)}")
        print(f"{'='*60}")
        
        if not os.path.exists(image_path):
            print(f"❌ Файл не знайдено: {image_path}")
            return None
        
        # Розпізнавання тексту
        print("🔍 Розпізнавання тексту...")
        text = self.extract_text_from_image(image_path)
        
        if text:
            print(f"\n📄 Розпізнаний текст:\n{'-'*60}\n{text}\n{'-'*60}")
        
        # Витягування даних
        print("\n🔎 Витягування даних...")
        data = self.parse_meter_data(text, image_path)
        
        # Можливість ручного введення або коригування
        if manual_data:
            data.update(manual_data)
        
        # Виведення знайдених даних
        print(f"\n📊 Знайдені дані:")
        print(f"  Марка: {data['brand'] or '(не знайдено)'}")
        print(f"  Серійний номер: {data['serial_number'] or '(не знайдено)'}")
        print(f"  Рік випуску: {data['year'] or '(не знайдено)'}")
        print(f"  Показання: {data['reading'] or '(не знайдено)'}")
        
        # Збереження в Excel
        print(f"\n💾 Збереження в Excel...")
        self.add_to_excel(data)
        
        return data


def main():
    """Приклад використання"""
    processor = MeterProcessor('lichilnyky.xlsx')
    
    # Приклад обробки одного фото
    image_path = input("Введіть шлях до фото лічильника: ").strip()
    
    if image_path:
        result = processor.process_image(image_path)
        
        if result:
            print(f"\n{'='*60}")
            print("✅ Обробку завершено!")
            print(f"{'='*60}")
            print(f"\nДані збережено в файл: {processor.excel_file}")
    else:
        print("❌ Шлях до фото не вказано")


if __name__ == '__main__':
    main()
